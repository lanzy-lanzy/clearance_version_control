import json
import logging
from datetime import datetime
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.files.storage import default_storage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView, ListView, DetailView

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Local imports
from core.models import (
    ClearanceRequest, Clearance, Staff, Student,
    Office, ProgramChair, User, Dean, Course,
    UserProfile, SEMESTER_CHOICES
)

# Set up logging
logger = logging.getLogger(__name__)

# Utility Functions
def get_current_school_year():
    current_date = datetime.now()
    return (f"{current_date.year - 1}-{current_date.year}"
            if current_date.month <= 5
            else f"{current_date.year}-{current_date.year + 1}")

def get_current_semester():
    current_month = datetime.now().month
    if 1 <= current_month <= 5:
        return "Second"
    elif 6 <= current_month <= 7:
        return "Summer"
    return "First"

def get_school_years():
    current_year = timezone.now().year
    return [f"{year}-{year + 1}" for year in range(current_year - 2, current_year + 1)]

def is_program_chair(user):
    return hasattr(user, 'programchair')

# Authentication Views
def user_login(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'student'):
            return redirect('student_dashboard')
        elif hasattr(request.user, 'programchair'):
            return redirect('program_chair_dashboard')
        elif hasattr(request.user, 'staff'):
            return redirect('staff_dashboard')
        elif request.user.is_superuser:
            return redirect('admin_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user:
            if not user.is_active:
                messages.error(request, 'Your account is not yet approved.')
                return redirect('login')

            login(request, user)
            if hasattr(user, 'student'):
                return redirect('student_dashboard')
            elif hasattr(user, 'programchair'):
                return redirect('program_chair_dashboard')
            elif hasattr(user, 'staff'):
                return redirect('staff_dashboard')
            elif user.is_superuser:
                return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid credentials.')

    return render(request, 'registration/login.html')

def user_logout(request):
    logout(request)
    return redirect('login')

def register(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        try:
            # Debug logging
            course_id = request.POST.get('course')
            print(f"Received course ID: {course_id}")
            print(f"All POST data: {request.POST}")

            username = request.POST.get('username')
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists.')
                return redirect('register')

            email = request.POST.get('email')
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists.')
                return redirect('register')

            user = User.objects.create_user(
                username=username,
                email=email,
                password=request.POST.get('password'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                is_active=False
            )

            try:
                course = Course.objects.get(id=course_id)
                print(f"Found course: {course.code} - {course.name}")
            except Course.DoesNotExist:
                print(f"No course found with ID: {course_id}")
                available_courses = Course.objects.all()
                print(f"Available course IDs: {[c.id for c in available_courses]}")
                raise Exception(f"Course with ID {course_id} does not exist")

            # Get the program chair for the selected course's dean
            program_chair = ProgramChair.objects.filter(dean=course.dean).first()

            # Get dormitory owner if student is a boarder
            dormitory_owner = None
            if request.POST.get('is_boarder') == 'on' and request.POST.get('dormitory_owner'):
                dormitory_owner = Staff.objects.filter(id=request.POST.get('dormitory_owner')).first()

            Student.objects.create(
                user=user,
                student_id=request.POST.get('student_id'),
                course=course,
                program_chair=program_chair,
                dormitory_owner=dormitory_owner,
                year_level=request.POST.get('year_level'),
                contact_number=request.POST.get('contact_number'),
                is_boarder=request.POST.get('is_boarder') == 'on',
                is_approved=False
            )

            messages.success(request, 'Registration successful! Awaiting admin approval.')
            return redirect('login')

        except Exception as e:
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Registration error: {str(e)}')

    context = {
        'deans': Dean.objects.all().order_by('name'),
        'dormitory_owners': Staff.objects.filter(is_dormitory_owner=True)
    }
    return render(request, 'registration/register.html', context)

# Basic Views
def home(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'programchair'):
            return redirect('program_chair_dashboard')
        elif hasattr(request.user, 'student'):
            return redirect('student_dashboard')
        elif hasattr(request.user, 'staff'):
            return redirect('staff_dashboard')
        elif request.user.is_superuser:
            return redirect('admin_dashboard')
    return render(request, 'home.html')

# Student Views
@login_required
def student_dashboard(request):
    try:
        student = request.user.student
        clearances = Clearance.objects.filter(student=student).order_by('-school_year', '-semester')

        # Get the latest clearance for the student
        latest_clearance = clearances.first()

        # Get all clearance requests for the student
        if latest_clearance:
            clearance_requests = ClearanceRequest.objects.filter(
                clearance=latest_clearance
            ).select_related('office', 'reviewed_by')
        else:
            clearance_requests = []

        # Get student's course and other related information
        course_info = student.course

        return render(request, 'core/student_dashboard.html', {
            'student_info': student,
            'clearances': clearances,
            'latest_clearance': latest_clearance,
            'clearance_requests': clearance_requests,
            'course_info': course_info,
        })
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('home')

@login_required
def student_profile(request):
    try:
        student = request.user.student
        student_info = {
            'full_name': student.user.get_full_name(),
            'student_id': student.student_id,
            'course': student.course.name,
            'year_level': student.year_level,
            'program_chair': student.course.dean.name,
            'is_boarder': student.is_boarder
        }
        return render(request, 'core/student_profile.html', {
            'student': student,
            'student_info': student_info
        })
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('home')

@login_required
def create_clearance_requests(request):
    if not hasattr(request.user, 'student'):
        messages.error(request, "Student profile required.")
        return redirect('home')

    student = request.user.student

    if request.method == 'POST':
        school_year = request.POST.get('school_year')
        semester = request.POST.get('semester')

        if not all([school_year, semester]):
            messages.error(request, "School year and semester required.")
            return redirect('create_clearance_requests')

        if Clearance.objects.filter(student=student, school_year=school_year, semester=semester).exists():
            messages.error(request, f"Clearance already exists for {school_year} {semester}.")
            return redirect('student_dashboard')

        if not student.is_approved:
            messages.error(request, "Account must be approved first.")
            return redirect('student_dashboard')

        try:
            clearance = Clearance.objects.create(
                student=student,
                school_year=school_year,
                semester=semester,
                is_cleared=False
            )

            required_offices = Office.objects.filter(
                Q(office_type='OTHER') | Q(office_type=student.course.dean.name)
            )

            for office in required_offices:
                ClearanceRequest.objects.create(
                    student=student,
                    clearance=clearance,
                    office=office,
                    school_year=school_year,
                    semester=semester,
                    status='pending'
                )

            messages.success(request, f"Clearance requests created for {school_year} {semester}.")
            return redirect('view_clearance_details', clearance_id=clearance.id)

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            logger.error(f"Error creating clearance requests: {str(e)}", exc_info=True)
            return redirect('student_dashboard')

    return render(request, 'core/create_clearance_requests.html', {
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
        'student': student
    })

@login_required
@require_POST
def request_again(request, request_id):
    clearance_request = get_object_or_404(ClearanceRequest, id=request_id)

    if clearance_request.student.user != request.user:
        messages.error(request, "Permission denied.")
        return redirect('student_dashboard')

    if clearance_request.status != 'denied':
        messages.error(request, "Can only request again for denied clearances.")
        return redirect('view_clearance_details', clearance_id=clearance_request.clearance.id)

    try:
        clearance_request.status = 'pending'
        clearance_request.reviewed_date = None
        clearance_request.reviewed_by = None
        clearance_request.notes = None
        clearance_request.request_date = timezone.now()
        clearance_request.save()

        messages.success(request, f"Resubmitted clearance request for {clearance_request.office.name}")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('view_clearance_details', clearance_id=clearance_request.clearance.id)

# Program Chair Views
@login_required
@user_passes_test(is_program_chair)
def program_chair_dashboard(request):
    program_chair = request.user.programchair
    students = Student.objects.filter(course__dean=program_chair.dean)

    current_year = timezone.now().year
    school_year = f"{current_year}-{current_year + 1}"
    semester = "1ST"

    total_students = students.count()
    cleared_students = students.filter(
        clearances__is_cleared=True,
        clearances__school_year=school_year,
        clearances__semester=semester
    ).count()
    pending_clearances = students.filter(
        clearances__is_cleared=False,
        clearances__school_year=school_year,
        clearances__semester=semester
    ).count()

    paginator = Paginator(students, 10)
    page = request.GET.get('page')
    students_page = paginator.get_page(page)

    return render(request, 'core/program_chair_dashboard.html', {
        'program_chair': program_chair,
        'students': students_page,
        'total_students': total_students,
        'cleared_students': cleared_students,
        'pending_clearances': pending_clearances,
        'school_year': school_year,
        'semester': semester,
        'page_obj': students_page,
    })

@login_required
@user_passes_test(is_program_chair)
def program_chair_profile(request):
    try:
        program_chair = request.user.programchair
        return render(request, 'core/program_chair_profile.html', {
            'program_chair': program_chair
        })
    except ProgramChair.DoesNotExist:
        messages.error(request, "Program chair profile not found.")
        return redirect('home')

@login_required
def delete_clearance(request, clearance_id):
    clearance = get_object_or_404(Clearance, id=clearance_id)
    if request.method == 'POST':
        clearance.delete()
        messages.success(request, 'Clearance deleted successfully.')
    return redirect('program_chair_dashboard')

class ManageStudentsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Student
    template_name = 'core/manage_students.html'
    context_object_name = 'students'
    paginate_by = 10

    def test_func(self):
        return hasattr(self.request.user, 'programchair')

    def get_queryset(self):
        program_chair = self.request.user.programchair
        queryset = Student.objects.filter(
            course__dean=program_chair.dean
        ).select_related('user', 'course').prefetch_related('clearances')

        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(student_id__icontains=search_query)
            )

        year_level = self.request.GET.get('year_level')
        if year_level:
            queryset = queryset.filter(year_level=year_level)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_year = timezone.now().year
        context.update({
            'program_chair': self.request.user.programchair,
            'current_school_year': f"{current_year}-{current_year + 1}",
            'current_semester': "1ST",
        })
        return context

    def handle_no_permission(self):
        messages.error(self.request, "Permission denied.")
        return redirect('home')

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        student_id = request.POST.get('student_id')

        try:
            student = Student.objects.get(id=student_id)
            if action == 'delete':
                student.user.delete()
                messages.success(request, 'Student deleted successfully.')
            elif action == 'approve':
                student.is_approved = True
                student.approval_date = timezone.now()
                student.approval_admin = request.user
                student.save()
                messages.success(request, 'Student approved successfully.')
        except Student.DoesNotExist:
            messages.error(request, 'Student not found.')

        return redirect('manage_students')

class StudentDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Student
    template_name = 'core/student_detail.html'
    context_object_name = 'student'

    def test_func(self):
        return hasattr(self.request.user, 'programchair')

    def get_queryset(self):
        return Student.objects.filter(
            course__dean=self.request.user.programchair.dean
        ).select_related('user', 'course')  # Remove 'program_chair' and 'program_chair__user'

    def handle_no_permission(self):
        messages.error(self.request, "Permission denied.")
        return redirect('home')

# Staff Views
@login_required
def staff_dashboard(request):
    staff = request.user.staff

    # Get current school year and semester
    current_year = timezone.now().year
    school_year = f"{current_year}-{current_year + 1}"
    semester = get_current_semester()

    # Get recent requests for this office
    recent_requests = ClearanceRequest.objects.filter(
        office=staff.office
    ).select_related(
        'student__user',
        'student__course'
    ).order_by('-request_date')[:10]  # Get last 10 requests using request_date

    # Get pending requests count
    pending_requests_count = ClearanceRequest.objects.filter(
        office=staff.office,
        status='pending'
    ).count()

    # Get approved requests count for today
    today = timezone.now().date()
    approved_today_count = ClearanceRequest.objects.filter(
        office=staff.office,
        status='approved',
        reviewed_date__date=today
    ).count()

    return render(request, 'core/staff_dashboard.html', {
        'recent_requests': recent_requests,
        'pending_requests_count': pending_requests_count,
        'approved_today_count': approved_today_count,
        'school_year': school_year,
        'current_semester': semester,
        'office': staff.office,
    })

@login_required
def staff_profile(request):
    try:
        staff = request.user.staff
        if request.method == 'POST' and 'update_role' in request.POST:
            staff.role = request.POST.get('role')
            staff.save()
            messages.success(request, 'Role updated successfully')
            return redirect('staff_profile')

        return render(request, 'core/staff_profile.html', {
            'staff': staff,
            'office': staff.office,
            'user': request.user,
            'is_dormitory_owner': staff.is_dormitory_owner,
            'assigned_students': staff.students_dorm.all() if staff.is_dormitory_owner else None,
        })
    except Staff.DoesNotExist:
        messages.error(request, "Staff profile not found.")
        return redirect('home')

@login_required
def staff_pending_requests(request):
    staff = request.user.staff
    pending_requests = ClearanceRequest.objects.filter(
        office=staff.office, status='pending'
    )

    search_query = request.GET.get('search', '')
    if search_query:
        pending_requests = pending_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    school_year = request.GET.get('school_year', '')
    if school_year:
        pending_requests = pending_requests.filter(school_year=school_year)

    semester = request.GET.get('semester', '')
    if semester:
        pending_requests = pending_requests.filter(semester=semester)

    sort_by = request.GET.get('sort', 'request_date')
    if sort_by == 'student_name':
        pending_requests = pending_requests.order_by('student__user__last_name', 'student__user__first_name')
    elif sort_by == 'course':
        pending_requests = pending_requests.order_by('student__course__code')
    else:
        pending_requests = pending_requests.order_by('-request_date')

    return render(request, 'core/staff_pending_requests.html', {
        'pending_requests': pending_requests,
        'school_years': ClearanceRequest.objects.values_list('school_year', flat=True).distinct(),
        'SEMESTER_CHOICES': SEMESTER_CHOICES,
        'current_filters': {'search': search_query, 'school_year': school_year, 'semester': semester, 'sort': sort_by},
        'office': staff.office,
        'school_year': get_current_school_year(),
        'current_semester': get_current_semester(),
    })

@login_required
@require_POST
def approve_clearance_request(request, request_id):
    try:
        # Get the staff member and clearance request
        staff = request.user.staff
        clearance_request = get_object_or_404(ClearanceRequest, id=request_id)

        # Log the request for debugging
        logger.info(f"Approve request received for request_id={request_id} by staff={staff.id}")

        # Check if staff has permission for this office
        if staff.office != clearance_request.office:
            messages.error(request, f"No permission for {clearance_request.office.name}")
            return redirect('staff_pending_requests')

        # Check if request is still pending
        if clearance_request.status != 'pending':
            messages.warning(request, f'Request already processed (current status: {clearance_request.status})')
            return redirect('staff_pending_requests')

        # Process the approval
        try:
            # Update the clearance request status
            clearance_request.status = "approved"
            clearance_request.reviewed_by = staff
            clearance_request.reviewed_date = timezone.now()
            clearance_request.notes = None  # Clear notes on approval
            clearance_request.save()

            # Update the parent clearance
            clearance = Clearance.objects.get(
                student=clearance_request.student,
                school_year=clearance_request.school_year,
                semester=clearance_request.semester
            )
            clearance.check_clearance()

            logger.info(f"Successfully approved request_id={request_id}")

            # Show success message and redirect
            messages.success(request, f'Successfully approved clearance for {clearance_request.student.user.get_full_name()}')
            return redirect('staff_pending_requests')

        except Exception as inner_e:
            logger.error(f"Error in approval process: {str(inner_e)}")
            messages.error(request, f'Error processing approval: {str(inner_e)}')
            return redirect('staff_pending_requests')

    except Staff.DoesNotExist:
        logger.error(f"Staff access required for user_id={request.user.id}")
        messages.error(request, 'Staff access required')
        return redirect('login')

    except Exception as e:
        logger.error(f"Unexpected error in approve_clearance_request: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
        return redirect('staff_pending_requests')

@login_required
@require_POST
def deny_clearance_request(request, request_id):
    try:
        # Get the staff member and clearance request
        staff = request.user.staff
        clearance_request = get_object_or_404(ClearanceRequest, id=request_id)

        # Log the request for debugging
        logger.info(f"Deny request received for request_id={request_id} by staff={staff.id}")

        # Check if staff has permission for this office
        if staff.office != clearance_request.office:
            messages.error(request, f"No permission for {clearance_request.office.name}")
            return redirect('staff_pending_requests')

        # Check if request is still pending
        if clearance_request.status != 'pending':
            messages.warning(request, f'Request already processed (current status: {clearance_request.status})')
            return redirect('staff_pending_requests')

        # Get the reason from form data
        reason = request.POST.get('remarks', '')

        if not reason:
            messages.error(request, 'Reason required for denial')
            return redirect('staff_pending_requests')

        try:
            # Update the clearance request status
            clearance_request.status = "denied"
            clearance_request.reviewed_by = staff
            clearance_request.reviewed_date = timezone.now()
            clearance_request.notes = reason
            clearance_request.save()

            # Update the parent clearance
            clearance = Clearance.objects.get(
                student=clearance_request.student,
                school_year=clearance_request.school_year,
                semester=clearance_request.semester
            )
            clearance.check_clearance()

            logger.info(f"Successfully denied request_id={request_id}")

            # Show success message and redirect
            messages.success(request, f'Successfully denied clearance for {clearance_request.student.user.get_full_name()}')
            return redirect('staff_pending_requests')

        except Exception as inner_e:
            logger.error(f"Error in denial process: {str(inner_e)}")
            messages.error(request, f'Error processing denial: {str(inner_e)}')
            return redirect('staff_pending_requests')

    except Staff.DoesNotExist:
        logger.error(f"Staff access required for user_id={request.user.id}")
        messages.error(request, 'Staff access required')
        return redirect('login')

    except Exception as e:
        logger.error(f"Unexpected error in deny_clearance_request: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
        return redirect('staff_pending_requests')

@login_required
def staff_clearance_history(request):
    try:
        staff = request.user.staff
    except Staff.DoesNotExist:
        return redirect('login')

    clearance_requests = ClearanceRequest.objects.filter(
        office=staff.office,
        status__in=['approved', 'denied']
    ).select_related('student', 'student__user', 'student__course', 'reviewed_by').order_by('-reviewed_date')

    status = request.GET.get('status', '')
    if status:
        clearance_requests = clearance_requests.filter(status=status)

    school_year = request.GET.get('school_year', '')
    if school_year:
        clearance_requests = clearance_requests.filter(clearance__school_year=school_year)

    semester = request.GET.get('semester', '')
    if semester:
        clearance_requests = clearance_requests.filter(clearance__semester=semester)

    search_query = request.GET.get('search', '')
    if search_query:
        clearance_requests = clearance_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    paginator = Paginator(clearance_requests, 10)
    page = request.GET.get('page')
    try:
        clearance_requests = paginator.page(page)
    except PageNotAnInteger:
        clearance_requests = paginator.page(1)
    except EmptyPage:
        clearance_requests = paginator.page(paginator.num_pages)

    return render(request, 'core/staff_clearance_history.html', {
        'clearance_requests': clearance_requests,
        'school_years': get_school_years(),
        'current_filters': {'status': status, 'school_year': school_year, 'semester': semester, 'search': search_query},
        'office': staff.office,
        'SEMESTER_CHOICES': SEMESTER_CHOICES,
    })

@login_required
def view_request(request, request_id):
    try:
        staff = request.user.staff
        request_obj = get_object_or_404(ClearanceRequest, id=request_id, office=staff.office)
        return render(request, 'core/staff_view_request.html', {'request_obj': request_obj})
    except Staff.DoesNotExist:
        messages.error(request, "Staff profile not found.")
        return redirect('home')

# Admin Views

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_program_chairs(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                user = User.objects.get(id=request.POST.get('user'))
                dean = None
                if request.POST.get('dean'):
                    dean = Dean.objects.get(id=request.POST.get('dean'))

                # Check if user already has a program chair profile
                if hasattr(user, 'programchair'):
                    messages.error(request, f'User {user.get_full_name()} is already a program chair.')
                else:
                    # Create program chair
                    program_chair = ProgramChair.objects.create(
                        user=user,
                        dean=dean,
                        designation=request.POST.get('designation')
                    )

                    # Create staff entry if it doesn't exist
                    office, _ = Office.objects.get_or_create(name="Program Chair", defaults={"description": "Handles final clearance approval"})
                    if not hasattr(user, 'staff'):
                        Staff.objects.create(user=user, office=office, role="Program Chair")

                    messages.success(request, f'Program Chair {program_chair.get_full_name()} added successfully.')

            elif action == 'edit':
                program_chair = ProgramChair.objects.get(id=request.POST.get('program_chair_id'))

                # Update dean
                if request.POST.get('dean'):
                    program_chair.dean = Dean.objects.get(id=request.POST.get('dean'))
                else:
                    program_chair.dean = None

                # Update designation
                program_chair.designation = request.POST.get('designation')
                program_chair.save()

                messages.success(request, f'Program Chair {program_chair.get_full_name()} updated successfully.')

            elif action == 'delete':
                program_chair = ProgramChair.objects.get(id=request.POST.get('program_chair_id'))
                delete_type = request.POST.get('delete_type', 'safe')

                if delete_type == 'safe':
                    # Check if there are any students associated with this program chair
                    if Student.objects.filter(program_chair=program_chair).exists():
                        messages.error(
                            request,
                            f"Cannot safely delete program chair '{program_chair.get_full_name()}' because they have associated students. "
                            f"Please reassign the students first or use Force Delete."
                        )
                    else:
                        # Safe to delete
                        program_chair_name = program_chair.get_full_name()
                        program_chair.delete()
                        messages.success(request, f'Program Chair {program_chair_name} safely deleted.')

                elif delete_type == 'force':
                    # Force delete - will remove program chair association from students
                    student_count = Student.objects.filter(program_chair=program_chair).count()

                    # Use a transaction to ensure all operations succeed or fail together
                    from django.db import transaction
                    with transaction.atomic():
                        # Remove program chair association from students
                        Student.objects.filter(program_chair=program_chair).update(program_chair=None)

                        # Delete the program chair
                        program_chair_name = program_chair.get_full_name()
                        program_chair.delete()

                    messages.success(
                        request,
                        f'Program Chair {program_chair_name} force deleted and {student_count} student(s) reassigned.'
                    )

        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        except Dean.DoesNotExist:
            messages.error(request, 'Dean not found.')
        except ProgramChair.DoesNotExist:
            messages.error(request, 'Program Chair not found.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    # Get all program chairs
    program_chairs = ProgramChair.objects.all().select_related('user', 'dean')

    # Add student count to each program chair
    for pc in program_chairs:
        pc.student_count = Student.objects.filter(program_chair=pc).count()

    # Get available users (those who are not already program chairs)
    program_chair_user_ids = program_chairs.values_list('user_id', flat=True)
    available_users = User.objects.exclude(id__in=program_chair_user_ids).filter(is_active=True)

    return render(request, 'admin/program_chairs.html', {
        'program_chairs': program_chairs,
        'available_users': available_users,
        'deans': Dean.objects.all()
    })
@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_reassign_students(request):
    selected_course_id = request.GET.get('course_to_delete') or None
    course = None
    students = None

    # If course_to_delete is provided in the URL, load that course
    if selected_course_id:
        try:
            course = Course.objects.get(id=selected_course_id)
            students = Student.objects.filter(course=course)
        except Course.DoesNotExist:
            messages.error(request, 'Course not found.')
            selected_course_id = None

    if request.method == 'POST':
        course_to_delete_id = request.POST.get('course_to_delete')
        course_to_reassign_id = request.POST.get('course_to_reassign')

        if course_to_delete_id and course_to_reassign_id:
            try:
                course_to_delete = Course.objects.get(id=course_to_delete_id)
                course_to_reassign = Course.objects.get(id=course_to_reassign_id)

                # Get all students in the course to delete
                students_to_reassign = Student.objects.filter(course=course_to_delete)
                count = students_to_reassign.count()

                if count > 0:
                    # Reassign all students to the new course
                    for student in students_to_reassign:
                        student.course = course_to_reassign
                        student.save()

                    # Now delete the course
                    course_to_delete.delete()

                    messages.success(
                        request,
                        f'Successfully reassigned {count} students from {course_to_delete.code} to {course_to_reassign.code} and deleted the course.'
                    )
                    return redirect('admin_courses')
                else:
                    # No students to reassign, just delete the course
                    course_to_delete.delete()
                    messages.success(request, f'Course {course_to_delete.code} deleted successfully.')
                    return redirect('admin_courses')

            except Course.DoesNotExist:
                messages.error(request, 'One or both courses not found.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        elif course_to_delete_id:
            # Just showing the students for the selected course
            selected_course_id = course_to_delete_id
            try:
                course = Course.objects.get(id=selected_course_id)
                students = Student.objects.filter(course=course)
            except Course.DoesNotExist:
                messages.error(request, 'Course not found.')

    # Get all courses with student count
    courses = Course.objects.annotate(student_count=Count('student')).all()

    return render(request, 'admin/reassign_students.html', {
        'courses': courses,
        'selected_course_id': selected_course_id,
        'course': course,
        'students': students,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    pending_approvals = Student.objects.filter(
        is_approved=False,
        user__is_active=False
    ).select_related('user', 'course').order_by('-user__date_joined')

    return render(request, 'admin/dashboard.html', {
        'total_students': Student.objects.count(),
        'total_staff': Staff.objects.count(),
        'total_program_chairs': ProgramChair.objects.count(),
        'clearance_stats': {
            'total': Clearance.objects.count(),
            'pending': ClearanceRequest.objects.filter(status='pending').count(),
            'approved': ClearanceRequest.objects.filter(status='approved').count(),
            'denied': ClearanceRequest.objects.filter(status='denied').count(),
        },
        'recent_clearances': Clearance.objects.select_related('student')[:5],
        'offices': Office.objects.annotate(
            staff_count=Count('staff'),
            pending_requests=Count('clearance_requests', filter=Q(clearance_requests__status='pending'))
        ),
        'pending_approvals': pending_approvals,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_profile(request):
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    return render(request, 'core/admin_profile.html', {
        'admin_user': request.user,
        'profile': profile
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_students(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        student_id = request.POST.get('student_id')

        if action == 'approve':
            try:
                student = Student.objects.get(id=student_id)
                student.is_approved = True
                student.approval_date = timezone.now()
                student.approval_admin = request.user
                student.user.is_active = True
                student.save()
                student.user.save()
                messages.success(request, f'Student {student.user.get_full_name()} approved successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')

        elif action == 'delete':
            try:
                student = Student.objects.get(id=student_id)
                student_name = student.user.get_full_name()
                # Delete the user which will cascade delete the student
                student.user.delete()
                messages.success(request, f'Student {student_name} deleted successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')

        elif action == 'add_student_form':
            return render(request, 'admin/add_student.html', {
                'courses': Course.objects.all(),
                'deans': Dean.objects.all().order_by('name'),
            })

        elif action == 'export_pdf':
            from core.pdf_utils import generate_students_pdf

            # Get the filtered students based on the current filters
            search_query = request.POST.get('search_query', '')
            course_filter = request.POST.get('course_filter', '')
            year_level_filter = request.POST.get('year_level_filter', '')
            status_filter = request.POST.get('status_filter', '')
            boarder_filter = request.POST.get('boarder_filter', '')

            # Start with all students
            students = Student.objects.select_related('user', 'course').all()

            # Apply filters
            if search_query:
                students = students.filter(
                    Q(user__first_name__icontains=search_query) |
                    Q(user__last_name__icontains=search_query) |
                    Q(student_id__icontains=search_query) |
                    Q(user__email__icontains=search_query)
                )

            if course_filter:
                students = students.filter(course_id=course_filter)

            if year_level_filter:
                students = students.filter(year_level=year_level_filter)

            if status_filter == 'active':
                students = students.filter(user__is_active=True)
            elif status_filter == 'inactive':
                students = students.filter(user__is_active=False)
            elif status_filter == 'pending':
                students = students.filter(is_approved=False)

            if boarder_filter == '1':
                students = students.filter(is_boarder=True)
            elif boarder_filter == '0':
                students = students.filter(is_boarder=False)

            # Generate PDF
            pdf = generate_students_pdf(students, request)

            # Create the HttpResponse with PDF headers
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="students_report.pdf"'

            return response

        elif action == 'view_details':
            try:
                student = Student.objects.get(id=student_id)
                return render(request, 'admin/view_student.html', {
                    'student': student,
                })
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
                return redirect('admin_students')

        elif action == 'edit_form':
            try:
                student = Student.objects.get(id=student_id)
                return render(request, 'admin/edit_student.html', {
                    'student': student,
                    'courses': Course.objects.all(),
                })
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
                return redirect('admin_students')

        elif action == 'edit_student':
            try:
                student = Student.objects.get(id=student_id)
                user = student.user

                # Update user information
                user.first_name = request.POST.get('first_name')
                user.last_name = request.POST.get('last_name')
                user.email = request.POST.get('email')
                user.save()

                # Update student information
                student.contact_number = request.POST.get('contact_number')
                if request.POST.get('course'):
                    student.course_id = request.POST.get('course')
                if request.POST.get('year_level'):
                    student.year_level = request.POST.get('year_level')
                student.is_boarder = request.POST.get('is_boarder') == 'on'
                student.is_approved = request.POST.get('is_approved') == 'on'
                student.user.is_active = student.is_approved
                student.save()
                user.save()

                messages.success(request, f'Student {student.user.get_full_name()} updated successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
            except Exception as e:
                messages.error(request, f'Error updating student: {str(e)}')

        elif action == 'add_student':
            try:
                # Create user
                user = User.objects.create_user(
                    username=request.POST.get('email').split('@')[0],  # Use first part of email as username
                    password='student123',  # Default password
                    email=request.POST.get('email'),
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    is_active=request.POST.get('auto_approve') == 'on'
                )

                # Generate student ID
                import random
                import string
                from django.utils.text import slugify

                year = timezone.now().year % 100
                first_chars = slugify(request.POST.get('first_name'))[:2].upper()
                last_chars = slugify(request.POST.get('last_name'))[:2].upper()
                random_digits = ''.join(random.choices(string.digits, k=4))
                student_id = f"{year}-{first_chars}{last_chars}-{random_digits}"

                # Get the course
                course_id = request.POST.get('course')

                # Get program chair if selected
                program_chair_id = request.POST.get('program_chair')
                program_chair = None
                if program_chair_id:
                    try:
                        program_chair = ProgramChair.objects.get(id=program_chair_id)
                    except ProgramChair.DoesNotExist:
                        pass

                # Create student
                student = Student.objects.create(
                    user=user,
                    student_id=student_id,
                    course_id=course_id,
                    program_chair=program_chair,
                    year_level=request.POST.get('year_level'),
                    contact_number=request.POST.get('contact_number'),
                    is_boarder=request.POST.get('is_boarder') == 'on',
                    is_approved=request.POST.get('auto_approve') == 'on'
                )

                # If no program chair was selected, try to assign one automatically
                if not program_chair:
                    course = Course.objects.get(id=course_id)
                    if course and course.dean:
                        # Try to find a program chair for this dean
                        auto_program_chair = ProgramChair.objects.filter(dean=course.dean).first()
                        if auto_program_chair:
                            student.program_chair = auto_program_chair
                            student.save()

                messages.success(request, f'Student {student.user.get_full_name()} added successfully with ID {student_id}.')
            except Exception as e:
                messages.error(request, f'Error adding student: {str(e)}')
                # Clean up if user was created but student wasn't
                if 'user' in locals() and 'student' not in locals():
                    user.delete()

    # Get query parameters for filtering
    search_query = request.GET.get('search', '')
    course_filter = request.GET.get('course', '')
    year_level_filter = request.GET.get('year_level', '')
    status_filter = request.GET.get('status', '')
    boarder_filter = request.GET.get('boarder', '')

    # Start with all students
    students = Student.objects.select_related('user', 'course').all()

    # Apply filters
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    if course_filter:
        students = students.filter(course_id=course_filter)

    if year_level_filter:
        students = students.filter(year_level=year_level_filter)

    if status_filter == 'active':
        students = students.filter(user__is_active=True)
    elif status_filter == 'inactive':
        students = students.filter(user__is_active=False)
    elif status_filter == 'pending':
        students = students.filter(is_approved=False)

    if boarder_filter == '1':
        students = students.filter(is_boarder=True)
    elif boarder_filter == '0':
        students = students.filter(is_boarder=False)

    return render(request, 'admin/students.html', {
        'students': students,
        'courses': Course.objects.all(),
        'search_query': search_query,
        'course_filter': course_filter,
        'year_level_filter': year_level_filter,
        'status_filter': status_filter,
        'boarder_filter': boarder_filter,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_staff(request):
    staff = Staff.objects.select_related('user', 'office').all().order_by('office__name', 'user__first_name')

    return render(request, 'admin/staff.html', {
        'staff': staff,
        'statistics': {
            'total_staff': staff.count(),
            'active_staff': staff.filter(user__is_active=True).count(),
            'total_offices': Office.objects.count(),
            'staff_by_office': Office.objects.annotate(staff_count=Count('staff')),
        }
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_staff_add(request):
    if request.method == 'POST':
        try:
            user = User.objects.create_user(
                username=request.POST.get('username'),
                email=request.POST.get('email'),
                password=request.POST.get('password'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                is_active=True
            )

            Staff.objects.create(
                user=user,
                office_id=request.POST.get('office'),
                role=request.POST.get('role'),
                is_dormitory_owner=request.POST.get('is_dormitory_owner') == 'on'
            )
            messages.success(request, 'Staff member added successfully.')
            return redirect('admin_staff')
        except Exception as e:
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error: {str(e)}')

    return render(request, 'admin/staff_add.html', {'offices': Office.objects.all()})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_staff_edit(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)

    if request.method == 'POST':
        try:
            staff.user.username = request.POST.get('username')
            staff.user.email = request.POST.get('email')
            staff.user.first_name = request.POST.get('first_name')
            staff.user.last_name = request.POST.get('last_name')
            if request.POST.get('password'):
                staff.user.set_password(request.POST.get('password'))
            staff.user.is_active = request.POST.get('is_active') == 'on'
            staff.user.save()

            staff.office_id = request.POST.get('office')
            staff.role = request.POST.get('role')
            staff.is_dormitory_owner = request.POST.get('is_dormitory_owner') == 'on'
            if 'profile_picture' in request.FILES:
                staff.profile_picture = request.FILES['profile_picture']
            staff.save()

            messages.success(request, 'Staff member updated successfully.')
            return redirect('admin_staff')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return render(request, 'admin/staff_edit.html', {
        'staff_member': staff,
        'offices': Office.objects.all(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_deans(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                dean = Dean.objects.create(
                    name=request.POST.get('name'),
                    description=request.POST.get('description'),
                    logo=request.FILES.get('logo') if 'logo' in request.FILES else None
                )
                messages.success(request, f'Dean {dean.name} added.')
            elif action == 'edit':
                dean = Dean.objects.get(id=request.POST.get('dean_id'))
                dean.name = request.POST.get('name')
                dean.description = request.POST.get('description')
                if 'logo' in request.FILES:
                    if dean.logo:
                        dean.logo.delete(save=False)
                    dean.logo = request.FILES['logo']
                dean.save()
                messages.success(request, f'Dean {dean.name} updated.')
            elif action == 'delete':
                dean = Dean.objects.get(id=request.POST.get('dean_id'))
                delete_type = request.POST.get('delete_type', 'safe')

                if delete_type == 'safe':
                    # Check if there are any courses associated with this dean
                    if dean.courses.exists():
                        messages.error(
                            request,
                            f"Cannot safely delete dean '{dean.name}' because it has associated courses. "
                            f"Please reassign the courses first or use a different delete option."
                        )
                    else:
                        # Safe to delete
                        dean_name = dean.name
                        if dean.logo:
                            dean.logo.delete(save=False)
                        dean.delete()
                        messages.success(request, f'Dean {dean_name} safely deleted.')

                elif delete_type == 'force':
                    # Force delete - will cascade delete all courses
                    course_count = dean.courses.count()
                    student_count = Student.objects.filter(course__dean=dean).count()

                    # If there are students, we need to handle them first
                    if student_count > 0:
                        messages.error(
                            request,
                            f"Cannot force delete dean '{dean.name}' because there are {student_count} students "
                            f"enrolled in its courses. Please reassign or delete these students first."
                        )
                    else:
                        # No students, safe to delete courses and dean
                        dean_name = dean.name
                        if dean.logo:
                            dean.logo.delete(save=False)

                        # Use a transaction to ensure all operations succeed or fail together
                        from django.db import transaction
                        with transaction.atomic():
                            # Delete all courses first
                            dean.courses.all().delete()
                            # Then delete the dean
                            dean.delete()

                        messages.success(
                            request,
                            f'Dean {dean_name} force deleted along with {course_count} course(s).'
                        )

                elif delete_type == 'cascade':
                    # Cascade delete - will delete everything including students
                    course_count = dean.courses.count()
                    student_count = Student.objects.filter(course__dean=dean).count()

                    # Use a transaction to ensure all operations succeed or fail together
                    from django.db import transaction
                    with transaction.atomic():
                        dean_name = dean.name
                        if dean.logo:
                            dean.logo.delete(save=False)

                        # Delete all students associated with this dean's courses
                        Student.objects.filter(course__dean=dean).delete()

                        # Delete all courses
                        dean.courses.all().delete()

                        # Delete the dean
                        dean.delete()

                    messages.success(
                        request,
                        f'Dean {dean_name} permanently deleted along with {course_count} course(s) and {student_count} student(s).'
                    )
            elif action == 'reassign':
                # Redirect to the reassign view
                dean_id = request.POST.get('dean_id')
                return redirect('admin_reassign_courses', dean_id=dean_id)
        except Dean.DoesNotExist:
            messages.error(request, 'Dean not found.')

    deans = Dean.objects.all()

    # Get additional statistics for the template
    total_courses = Course.objects.count()
    latest_dean = deans.order_by('-created_at').first() if deans.exists() else None

    # Add course count to each dean
    for dean in deans:
        dean.course_count = dean.courses.count()
        dean.student_count = Student.objects.filter(course__dean=dean).count()

    return render(request, 'admin/deans.html', {
        'deans': deans,
        'total_courses': total_courses,
        'latest_dean': latest_dean
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_reassign_courses(request, dean_id):
    """View to reassign courses from one dean to another"""
    try:
        source_dean = Dean.objects.get(id=dean_id)
    except Dean.DoesNotExist:
        messages.error(request, "Dean not found.")
        return redirect('admin_deans')

    # Get all courses for this dean
    courses = source_dean.courses.all()
    if not courses.exists():
        messages.warning(request, f"No courses found for dean '{source_dean.name}'.")
        return redirect('admin_deans')

    # Add student count to each course for display
    for course in courses:
        course.student_count = Student.objects.filter(course=course).count()

    # Get all other deans
    available_deans = Dean.objects.exclude(id=dean_id)

    if request.method == 'POST':
        target_dean_id = request.POST.get('target_dean')
        if not target_dean_id:
            messages.error(request, "Please select a target dean.")
            return redirect('admin_reassign_courses', dean_id=dean_id)

        try:
            target_dean = Dean.objects.get(id=target_dean_id)

            # Use a transaction to ensure all courses are reassigned or none
            from django.db import transaction
            with transaction.atomic():
                count = courses.update(dean=target_dean)

            messages.success(
                request,
                f"Successfully reassigned {count} courses from '{source_dean.name}' to '{target_dean.name}'."
            )
            return redirect('admin_deans')
        except Exception as e:
            messages.error(request, f"Error reassigning courses: {str(e)}")
            return redirect('admin_reassign_courses', dean_id=dean_id)

    return render(request, 'admin/reassign_dean_courses.html', {
        'source_dean': source_dean,
        'available_deans': available_deans,
        'courses': courses,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_offices(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                office = Office.objects.create(
                    name=request.POST.get('name'),
                    description=request.POST.get('description'),
                    location=request.POST.get('location'),
                    contact_number=request.POST.get('contact_number'),
                    email=request.POST.get('email')
                )
                messages.success(request, f'Office "{office.name}" added.')
            elif action == 'edit':
                office = Office.objects.get(id=request.POST.get('office_id'))
                office.name = request.POST.get('name')
                office.description = request.POST.get('description')
                office.location = request.POST.get('location')
                office.contact_number = request.POST.get('contact_number')
                office.email = request.POST.get('email')
                office.save()
                messages.success(request, f'Office "{office.name}" updated.')
            elif action == 'delete':
                office = Office.objects.get(id=request.POST.get('office_id'))
                # Check if there are staff members associated with this office
                if office.staff.exists():
                    staff_list = ", ".join([staff.get_full_name() for staff in office.staff.all()[:5]])
                    if office.staff.count() > 5:
                        staff_list += f" and {office.staff.count() - 5} more"

                    messages.error(
                        request,
                        f"Cannot delete office '{office.name}' because it has associated staff members: {staff_list}. "
                        f"Please reassign or delete these staff members first."
                    )
                else:
                    # Safe to delete
                    office_name = office.name
                    office.delete()
                    messages.success(request, f'Office "{office_name}" deleted.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    offices = Office.objects.annotate(
        staff_count=Count('staff'),
        pending_requests=Count('clearance_requests', filter=Q(clearance_requests__status='pending')),
        completed_requests=Count('clearance_requests', filter=Q(clearance_requests__status='completed')),
        total_requests=Count('clearance_requests')
    ).order_by('name')

    return render(request, 'admin/offices.html', {
        'offices': offices,
        'total_offices': offices.count(),
        'total_staff': Staff.objects.count(),
        'total_pending_requests': sum(office.pending_requests for office in offices),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_courses(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                dean = get_object_or_404(Dean, id=request.POST.get('dean'))
                Course.objects.create(
                    code=request.POST.get('code'),
                    name=request.POST.get('name'),
                    dean=dean
                )
                messages.success(request, f'Course {request.POST.get("code")} added.')
            elif action == 'delete':
                course = get_object_or_404(Course, id=request.POST.get('course_id'))

                # Check if there are students associated with this course
                students = Student.objects.filter(course=course)
                if students.exists():
                    student_list = ", ".join([f"{student.get_full_name()} ({student.student_id})" for student in students[:5]])
                    if students.count() > 5:
                        student_list += f" and {students.count() - 5} more"

                    error_message = f"Cannot delete course '{course.code}' because it is referenced by students: {student_list}. " \
                                   f"Please reassign these students to another course first."
                    messages.error(request, error_message)
                else:
                    course.delete()
                    messages.success(request, f'Course {course.code} deleted.')
            elif action == 'edit':
                course = get_object_or_404(Course, id=request.POST.get('course_id'))
                course.code = request.POST.get('code')
                course.name = request.POST.get('name')
                course.dean = get_object_or_404(Dean, id=request.POST.get('dean'))
                course.save()
                messages.success(request, f'Course {course.code} updated.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('admin_courses')

    return render(request, 'admin/courses.html', {
        'courses': Course.objects.annotate(student_count=Count('student')).select_related('dean').all(),
        'deans': Dean.objects.all(),
        'total_students': Student.objects.count(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_clearances(request):
    clearances = Clearance.objects.select_related('student').all().order_by('-school_year', '-semester')
    return render(request, 'admin/clearances.html', {
        'clearances': clearances,
        'clearance_stats': {
            'pending': ClearanceRequest.objects.filter(status='PENDING').count(),
            'approved': ClearanceRequest.objects.filter(status='APPROVED').count(),
            'denied': ClearanceRequest.objects.filter(status='DENIED').count(),
        },
        'recent_clearances': Clearance.objects.select_related('student').order_by('-created_at')[:5],
    })
@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_pending_approvals(request):
    pending_students = Student.objects.filter(
        is_approved=False
    ).select_related('user', 'course')

    courses = Course.objects.all()

    return render(request, 'admin/pending_approvals.html', {
        'pending_students': pending_students,
        'total_pending': pending_students.count(),
        'courses': courses,
    })
@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_users(request):
    return render(request, 'admin/users.html', {
        'students': Student.objects.select_related('user', 'course').all(),
        'staff': Staff.objects.select_related('user', 'office').all(),
        'program_chairs': ProgramChair.objects.select_related('user').all(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_settings(request):
    if request.method == 'POST':
        # Handle settings updates
        pass

    return render(request, 'admin/settings.html', {
        'current_school_year': timezone.now().year,
        'current_semester': get_current_semester(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def create_user(request):
    if request.method == 'POST':
        try:
            user = User.objects.create_user(
                username=request.POST.get('username'),
                password=request.POST.get('password'),
                email=request.POST.get('email'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name')
            )

            user_type = request.POST.get('user_type')
            if user_type == 'student':
                Student.objects.create(
                    user=user,
                    student_id=request.POST.get('student_id'),
                    course=Course.objects.get(id=request.POST.get('course')),
                    program_chair=ProgramChair.objects.get(id=request.POST.get('program_chair')),
                    year_level=request.POST.get('year_level'),
                    is_approved=True,
                    approval_date=timezone.now(),
                    approval_admin=request.user
                )
            elif user_type == 'staff':
                Staff.objects.create(
                    user=user,
                    office=Office.objects.get(id=request.POST.get('office')),
                    is_dormitory_owner=request.POST.get('is_dormitory_owner') == 'on'
                )
            elif user_type == 'program_chair':
                ProgramChair.objects.create(
                    user=user,
                    dean=Dean.objects.get(id=request.POST.get('dean'))
                )

            messages.success(request, f'User {user.username} created.')
            return redirect('admin_users')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return render(request, 'admin/create_user.html', {
        'courses': Course.objects.all(),
        'offices': Office.objects.all(),
        'program_chairs': ProgramChair.objects.all(),
        'deans': Dean.objects.all(),
    })

# Clearance Views
@login_required
def update_clearance_request(request, request_id):
    clearance_request = get_object_or_404(ClearanceRequest, pk=request_id)

    try:
        staff_member = request.user.staff

        if staff_member.office != clearance_request.office:
            messages.error(request, f"No permission for {clearance_request.office.name}")
            return redirect('staff_dashboard')

        if clearance_request.office.name == "DORMITORY":
            if not staff_member.is_dormitory_owner or clearance_request.student.dormitory_owner != staff_member:
                messages.error(request, "Dormitory clearance permission denied.")
                return redirect('staff_dashboard')

        if clearance_request.office.name.startswith('SSB'):
            if clearance_request.office.affiliated_dean != clearance_request.student.course.dean:
                messages.error(request, "SSB clearance permission denied.")
                return redirect('staff_dashboard')

    except Staff.DoesNotExist:
        messages.error(request, "Staff access required.")
        return redirect('login')

    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '')

        try:
            if action == 'approve':
                clearance_request.approve(staff_member)
                messages.success(request, f"Approved for {clearance_request.student.full_name}")
            elif action == 'deny':
                if not remarks:
                    messages.error(request, "Remarks required for denial.")
                    return redirect('staff_dashboard')
                clearance_request.deny(staff_member, remarks)
                messages.success(request, f"Denied for {clearance_request.student.full_name}")
            else:
                messages.error(request, "Invalid action.")
                return redirect('staff_dashboard')

            clearance = Clearance.objects.get(
                student=clearance_request.student,
                school_year=clearance_request.school_year,
                semester=clearance_request.semester
            )
            clearance.check_clearance()

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return redirect('staff_dashboard')

@login_required
def view_clearance_details(request, clearance_id):
    clearance = get_object_or_404(Clearance, id=clearance_id)

    if request.user.is_superuser:
        pass
    elif hasattr(request.user, 'programchair'):
        if clearance.student.course.dean != request.user.programchair.dean:
            messages.error(request, "Permission denied.")
            return redirect('program_chair_dashboard')
    elif hasattr(request.user, 'staff'):
        if not ClearanceRequest.objects.filter(clearance=clearance, office=request.user.staff.office).exists():
            messages.error(request, "Permission denied.")
            return redirect('staff_dashboard')
    elif hasattr(request.user, 'student'):
        if request.user.student != clearance.student:
            messages.error(request, "Permission denied.")
            return redirect('student_dashboard')
    else:
        messages.error(request, "Permission denied.")
        return redirect('home')

    clearance_requests = ClearanceRequest.objects.filter(clearance=clearance).select_related('office', 'reviewed_by')

    return render(request, 'core/clearance_details.html', {
        'clearance': clearance,
        'clearance_requests': clearance_requests,
        'user_type': ('admin' if request.user.is_superuser else
                     'program_chair' if hasattr(request.user, 'programchair') else
                     'staff' if hasattr(request.user, 'staff') else 'student')
    })

# Report Views
@login_required
def generate_reports(request):
    if request.method == 'POST':
        # Handle report generation
        pass

    return render(request, 'core/generate_reports.html', {
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
    })

@login_required
def generate_report(request):
    if request.method == 'POST':
        school_year = request.POST.get('school_year')
        semester = request.POST.get('semester')
        report_type = request.POST.get('report_type')

        if report_type == 'pdf':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="clearance_report_{school_year}_{semester}.pdf"'
            # Add PDF generation logic
            return response
        elif report_type == 'excel':
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="clearance_report_{school_year}_{semester}.xlsx"'
            # Add Excel generation logic
            return response

        messages.success(request, 'Report generated successfully')
    return redirect('generate_reports')

@login_required
def export_clearances_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clearance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="44B78B", end_color="44B78B", fill_type="solid")

    headers = ['Student ID', 'Student Name', 'Course', 'Year Level', 'School Year', 'Semester',
              'Clearance Status', 'Date Cleared']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 15

    clearances = Clearance.objects.select_related('student', 'student__course').order_by(
        '-school_year', '-semester', 'student__student_id')

    for row, clearance in enumerate(clearances, 2):
        student = clearance.student
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.full_name)
        ws.cell(row=row, column=3, value=student.course.code)
        ws.cell(row=row, column=4, value=student.year_level)
        ws.cell(row=row, column=5, value=clearance.school_year)
        ws.cell(row=row, column=6, value=dict(SEMESTER_CHOICES)[clearance.semester])
        ws.cell(row=row, column=7, value="Cleared" if clearance.is_cleared else "Pending")
        ws.cell(row=row, column=8, value=clearance.cleared_date.strftime("%Y-%m-%d") if clearance.cleared_date else "N/A")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=clearance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

# Utility Views
@login_required
@require_POST
def update_profile_picture(request):
    try:
        if 'profile_picture' not in request.FILES:
            messages.error(request, 'No image provided')
            return redirect('home')

        if hasattr(request.user, 'student'):
            profile = request.user.student
            redirect_url = 'student_profile'
        elif hasattr(request.user, 'programchair'):
            profile = request.user.programchair
            redirect_url = 'program_chair_profile'
        elif hasattr(request.user, 'staff'):
            profile = request.user.staff
            redirect_url = 'staff_profile'
        elif request.user.is_superuser:
            profile, _ = UserProfile.objects.get_or_create(user=request.user)
            redirect_url = 'admin_profile'
        else:
            messages.error(request, 'Invalid user type')
            return redirect('home')

        if profile.profile_picture:
            profile.profile_picture.delete(save=False)
        profile.profile_picture = request.FILES['profile_picture']
        profile.save()

        messages.success(request, 'Profile picture updated')
        return redirect(redirect_url)
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('home')

@login_required
@require_POST
def update_contact_number(request):
    try:
        if not hasattr(request.user, 'student'):
            messages.error(request, 'Only students can update contact numbers')
            return redirect('home')

        student = request.user.student
        contact_number = request.POST.get('contact_number', '').strip()

        # Basic validation
        if contact_number and not contact_number.startswith('+'):
            contact_number = '+' + contact_number

        # Update the contact number
        student.contact_number = contact_number
        student.save()

        messages.success(request, 'Contact number updated successfully')
        return redirect('student_profile')
    except Exception as e:
        messages.error(request, f'Error updating contact number: {str(e)}')
        return redirect('student_profile')

@login_required
def print_permit(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'core/print_permit.html', {
        'student': student,
        'logo_url': 'images/logo.png',
    })

# API Endpoints
@login_required
def get_program_chairs(request, dean_id):
    program_chairs = ProgramChair.objects.filter(dean_id=dean_id).select_related('user')
    data = [{
        'id': pc.id,
        'user': {
            'full_name': f"{pc.user.get_full_name()}"
        }
    } for pc in program_chairs]
    return JsonResponse(data, safe=False)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_program_chair_details(request, program_chair_id):
    try:
        program_chair = ProgramChair.objects.get(id=program_chair_id)
        return JsonResponse({
            'id': program_chair.id,
            'user_id': program_chair.user.id,
            'full_name': program_chair.get_full_name(),
            'email': program_chair.user.email,
            'dean_id': program_chair.dean.id if program_chair.dean else None,
            'dean_name': program_chair.dean.name if program_chair.dean else None,
            'designation': program_chair.designation or '',
            'student_count': Student.objects.filter(program_chair=program_chair).count()
        })
    except ProgramChair.DoesNotExist:
        return JsonResponse({'error': 'Program Chair not found'}, status=404)

from django.http import JsonResponse
from django.core.exceptions import ObjectDoesNotExist
from .models import Course, Dean

def get_courses(request, dean_id):
    """
    Get courses for a specific dean. This endpoint is public and doesn't require authentication
    since it's used on the registration page.
    """
    try:
        # Debug logging
        print(f"Fetching courses for dean_id: {dean_id}")

        courses = Course.objects.filter(
            dean_id=dean_id,
            is_active=True
        ).order_by('code')

        # Debug logging
        print(f"Found courses: {[f'{c.code} (ID: {c.id})' for c in courses]}")

        data = [{
            'id': course.id,
            'code': course.code,
            'name': course.name
        } for course in courses]

        return JsonResponse(data, safe=False)

    except ObjectDoesNotExist:
        print(f"Dean not found with ID: {dean_id}")
        return JsonResponse(
            {'error': 'Dean not found'},
            status=404
        )
    except Exception as e:
        print(f"Error fetching courses: {str(e)}")
        return JsonResponse(
            {'error': str(e)},
            status=500
        )

@login_required
def get_offices(request, dean_id):
    offices = Office.objects.filter(affiliated_dean_id=dean_id)
    data = [{'id': office.id, 'name': office.name} for office in offices]
    return JsonResponse(data, safe=False)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_dean_details(request, dean_id):
    try:
        dean = Dean.objects.get(id=dean_id)
        return JsonResponse({
            'name': dean.name,
            'description': dean.description,
            'logo_url': dean.logo.url if dean.logo else None
        })
    except Dean.DoesNotExist:
        return JsonResponse({'error': 'Dean not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_user_details(request, user_id):
    try:
        student = Student.objects.select_related('user', 'course').get(user_id=user_id)
        return JsonResponse({
            'full_name': student.user.get_full_name(),
            'student_id': student.student_id,
            'course': student.course.name if student.course else 'N/A',
            'date_joined': student.user.date_joined.strftime('%B %d, %Y'),
            'email': student.user.email,
            'contact_number': student.contact_number or 'Not provided',
        })
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_student_details(request, student_id):
    try:
        student = Student.objects.select_related('user', 'course').get(id=student_id)
        return JsonResponse({
            'id': student.id,
            'full_name': f"{student.user.first_name} {student.user.last_name}",
            'first_name': student.user.first_name,
            'last_name': student.user.last_name,
            'student_id': student.student_id,
            'email': student.user.email,
            'contact_number': student.contact_number or '',
            'course': student.course.code if student.course else '',
            'course_id': student.course.id if student.course else '',
            'year_level': student.year_level,
            'date_applied': student.user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            'is_boarder': student.is_boarder,
            'is_approved': student.is_approved,
            'is_active': student.user.is_active,
            'profile_picture_url': student.get_profile_picture_url() if hasattr(student, 'get_profile_picture_url') else None,
            'created_at': student.created_at.strftime('%Y-%m-%d') if hasattr(student, 'created_at') and student.created_at else ''
        })
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_staff(request, staff_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        staff = Staff.objects.get(id=staff_id)
        user = staff.user
        staff.delete()
        user.delete()
        return JsonResponse({'success': True})
    except Staff.DoesNotExist:
        return JsonResponse({'error': 'Staff not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def approve_student(request, student_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        # Parse request body for notes
        data = json.loads(request.body)
        notes = data.get('notes', '')
        send_sms_notification = data.get('send_sms', True)  # Default to True

        student = Student.objects.get(id=student_id)
        student.approve_student(request.user)

        # If there are notes, you could store them in a new field or in a separate model
        # For now, we'll just log them
        if notes:
            logging.info(f"Approval notes for student {student.student_id}: {notes}")

            # You could add a field to store approval notes if needed:
            # student.approval_notes = notes
            # student.save()

        # Send SMS notification if requested and contact number exists
        sms_status = None
        if send_sms_notification and student.contact_number:
            from core.utils import send_sms

            # Prepare the message
            student_name = student.user.get_full_name()
            message = f"Hello {student_name}, your registration for {student.course.code} has been approved. You can now log in to the system."

            # Add notes if provided
            if notes:
                message += f" Note: {notes}"

            # Use the actual Twilio service
            # Set to False to make real API calls to Twilio
            use_test_mode = False

            # Send the SMS
            success, result, details = send_sms(student.contact_number, message, test_mode=use_test_mode)

            if success:
                sms_status = {
                    'sent': True,
                    'message_id': result,
                    'test_mode': use_test_mode
                }
                if use_test_mode:
                    logging.info(f"[TEST MODE] Approval SMS would be sent to student {student.student_id} at {student.contact_number}")
                else:
                    logging.info(f"Approval SMS sent to student {student.student_id} at {student.contact_number}")
            else:
                error_message = details.get('message', result) if isinstance(details, dict) else result
                sms_status = {
                    'sent': False,
                    'error': error_message,
                    'error_type': details.get('error_type') if isinstance(details, dict) else 'UNKNOWN_ERROR',
                    'error_code': details.get('error_code') if isinstance(details, dict) else None
                }
                logging.warning(f"Failed to send approval SMS to student {student.student_id}: {error_message}")
        elif send_sms_notification and not student.contact_number:
            sms_status = {'sent': False, 'error': 'No contact number available'}
            logging.warning(f"Cannot send approval SMS to student {student.student_id}: No contact number available")

        return JsonResponse({'success': True, 'sms_status': sms_status})
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    except Exception as e:
        logging.error(f"Error approving student {student_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def reject_student(request, student_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        data = json.loads(request.body)
        reason = data.get('reason')
        send_sms_notification = data.get('send_sms', True)  # Default to True

        if not reason:
            return JsonResponse({'success': False, 'error': 'Reason is required'}, status=400)

        student = Student.objects.get(id=student_id)
        student.reject_student(reason)

        # Log the rejection
        logging.info(f"Student {student.student_id} rejected. Reason: {reason}")

        # Send SMS notification if requested and contact number exists
        sms_status = None
        if send_sms_notification and student.contact_number:
            from core.utils import send_sms

            # Prepare the message
            student_name = student.user.get_full_name()
            message = f"Hello {student_name}, your registration for {student.course.code} has been rejected. Reason: {reason}"

            # Use the actual Twilio service
            # Set to False to make real API calls to Twilio
            use_test_mode = False

            # Send the SMS
            success, result, details = send_sms(student.contact_number, message, test_mode=use_test_mode)

            if success:
                sms_status = {
                    'sent': True,
                    'message_id': result,
                    'test_mode': use_test_mode
                }
                if use_test_mode:
                    logging.info(f"[TEST MODE] Rejection SMS would be sent to student {student.student_id} at {student.contact_number}")
                else:
                    logging.info(f"Rejection SMS sent to student {student.student_id} at {student.contact_number}")
            else:
                error_message = details.get('message', result) if isinstance(details, dict) else result
                sms_status = {
                    'sent': False,
                    'error': error_message,
                    'error_type': details.get('error_type') if isinstance(details, dict) else 'UNKNOWN_ERROR',
                    'error_code': details.get('error_code') if isinstance(details, dict) else None
                }
                logging.warning(f"Failed to send rejection SMS to student {student.student_id}: {error_message}")
        elif send_sms_notification and not student.contact_number:
            sms_status = {'sent': False, 'error': 'No contact number available'}
            logging.warning(f"Cannot send rejection SMS to student {student.student_id}: No contact number available")

        return JsonResponse({'success': True, 'sms_status': sms_status})
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    except Exception as e:
        logging.error(f"Error rejecting student {student_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def office_detail_api(request, office_id):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    office = get_object_or_404(Office, id=office_id)
    data = {'name': office.name}
    if hasattr(office, 'description'):
        data['description'] = office.description
    return JsonResponse(data)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_approval_stats(request):
    total_pending = Student.objects.filter(user__is_active=False).count()
    return JsonResponse({'total_pending': total_pending})

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Course, Dean  # Import your models

@require_http_methods(["GET"])
def get_courses_by_dean(request, dean_name):
    try:
        dean = Dean.objects.get(name=dean_name)
        courses = Course.objects.filter(dean=dean)
        data = [{'code': course.code, 'name': course.name} for course in courses]
        return JsonResponse(data, safe=False)
    except Dean.DoesNotExist:
        return JsonResponse([], safe=False)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def clearance_details(request, clearance_id):
    clearance = get_object_or_404(Clearance.objects.select_related(
        'student',
        'student__user',
        'student__course'
    ), id=clearance_id)

    clearance_requests = ClearanceRequest.objects.filter(
        clearance=clearance
    ).select_related('office', 'staff')

    return render(request, 'admin/clearance_details.html', {
        'clearance': clearance,
        'clearance_requests': clearance_requests,
    })











